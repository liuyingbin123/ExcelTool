from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, colors, Alignment, Border, borders


class ResultTable:
    """
    输出结果表
    """
    def __init__(self, path):
        self.wb = Workbook()
        self.path = path
        self.ws = self.wb.active

    def setTableHead(self, title, startrow=None, startcol=None, endrow=None, endcol=None):
        """
        :param title: 表名
        :param startrow: 合并单元格的起始行
        :param startcol: 合并单元格的起始列
        :param endrow: 合并单元格的结束行
        :param endcol: 合并单元格的结束行
        :return:
        """
        cell = self.ws.cell(row=1, column=1, value=title)
        # 20号字体，红色，居中对齐
        font = Font(name='宋体', size=20, color=colors.Color(rgb='00FF0000'))
        aligment = Alignment(horizontal='center', vertical='center')
        border = Border(left=borders.Side(style='medium', color='FF000000'),
                        right=borders.Side(style='medium', color='FF000000'),
                        top=borders.Side(style='medium', color='FF000000'),
                        bottom=borders.Side(style='medium', color='FF000000'),
                        diagonal=borders.Side(style='medium', color='FF000000'),
                        diagonal_direction=0,
                        outline=borders.Side(style='medium', color='FF000000'),
                        vertical=borders.Side(style='medium', color='FF000000'),
                        horizontal=borders.Side(style='medium', color='FF000000'))
        cell.font = font
        cell.alignment = aligment
        cell.border = border
        # 合并单元格
        if startrow and startcol and endrow and endcol:
            self.ws.merge_cells(start_row=startrow, start_column=startcol, end_row=endrow, end_column=endcol)

    def setData(self, row, col, data, startrow=None, startcol=None, endrow=None, endcol=None):
        cell = self.ws.cell(row=row, column=col, value=data)
        # 20号字体，红色，居中对齐
        font = Font(name='宋体', size=11, color=colors.BLACK)
        aligment = Alignment(horizontal='center', vertical='center')
        border = Border(left=borders.Side(style='medium', color='FF000000'),
                        right=borders.Side(style='medium', color='FF000000'),
                        top=borders.Side(style='medium', color='FF000000'),
                        bottom=borders.Side(style='medium', color='FF000000'),
                        diagonal=borders.Side(style='medium', color='FF000000'),
                        diagonal_direction=0,
                        outline=borders.Side(style='medium', color='FF000000'),
                        vertical=borders.Side(style='medium', color='FF000000'),
                        horizontal=borders.Side(style='medium', color='FF000000'))
        cell.font = font
        cell.alignment = aligment
        cell.border = border
        # 合并单元格
        if startrow and startcol and endrow and endcol:
            self.ws.merge_cells(start_row=startrow, start_column=startcol, end_row=endrow, end_column=endcol)

    def save(self):
        self.wb.save(self.path)